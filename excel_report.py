#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 17 22:16:26 2026

@author: simayarslan
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference


def baslik_formatla(ws, renk):
    doldur = PatternFill(start_color=renk, end_color=renk, fill_type="solid")
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.fill      = doldur
        cell.alignment = Alignment(horizontal="center")


def sutun_genislet(ws, genislik=22):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = genislik


def bar_grafik(ws, baslik, x_baslik, y_baslik, veri_col, kat_col, konum):
    g = BarChart()
    g.title        = baslik
    g.x_axis.title = x_baslik
    g.y_axis.title = y_baslik
    g.style        = 10
    satir = ws.max_row
    g.add_data(Reference(ws, min_col=veri_col, min_row=1, max_row=satir),
               titles_from_data=True)
    g.set_categories(Reference(ws, min_col=kat_col, min_row=2, max_row=satir))
    ws.add_chart(g, konum)


def excel_raporu_olustur(tablolar, csv_yol, excel_yol):
    # CSV kaydet
    writer = None
    try:
        df_ham = pd.read_csv(csv_yol, encoding="utf-8-sig")
        writer = pd.ExcelWriter(excel_yol, engine="openpyxl")
        df_ham.to_excel(writer, sheet_name="Tum Urunler", index=False)
        for sayfa_adi, tablo in tablolar.items():
            tablo.to_excel(writer, sheet_name=sayfa_adi, index=False)
        print("  [✓] Tablolar Excel'e yazildi.")
    except PermissionError:
        print(f"HATA: '{excel_yol}' baska bir program tarafindan acik.")
        raise
    except Exception as e:
        print(f"HATA: Excel yazilirken hata: {e}")
        raise
    finally:
        if writer is not None:
            writer.close()
            print("  (Sistem Mesaji: Excel baglantisi guvenle kapatildi.)")

    # Grafikleri ekledik
    try:
        wb = load_workbook(excel_yol)

        ws = wb["G1_Kat_Fiyat"]
        baslik_formatla(ws, "4472C4")
        sutun_genislet(ws)
        bar_grafik(ws, "Kategoriye Gore Ortalama Fiyat (TL)",
                   "Kategori", "Ort. Fiyat (TL)", 2, 1, "D2")
        print("  [✓] Grafik 1: Kategori bazli ortalama fiyat")

        ws = wb["G2_En_Yuksek_Puan"]
        baslik_formatla(ws, "FFC000")
        sutun_genislet(ws)
        bar_grafik(ws, "En Yuksek Puanli 10 Marka",
                   "Marka", "Ort. Puan", 2, 1, "D2")
        print("  [✓] Grafik 2: En yuksek puanli 10 marka")

        ws = wb["G3_Marka_Yorum"]
        baslik_formatla(ws, "9DC3E6")
        sutun_genislet(ws)
        bar_grafik(ws, "Marka Basina Ort. Yorum Sayisi (Top 10)",
                   "Marka", "Ort. Yorum Sayisi", 2, 1, "D2")
        print("  [✓] Grafik 3: Marka basina ortalama yorum sayisi")

        ws = wb["G4_Kat_Ozet"]
        baslik_formatla(ws, "C9C9FF")
        sutun_genislet(ws)
        print("  [✓] Tablo 4: Kategori ozet tablosu")

        ws = wb["G5_Fiyat_Segment"]
        baslik_formatla(ws, "FFD966")
        sutun_genislet(ws)
        bar_grafik(ws, "Fiyat Segmentine Gore Ortalama Puan",
                   "Segment", "Ort. Puan", 4, 1, "G2")
        bar_grafik(ws, "Fiyat Segmentine Gore Ortalama Yorum Sayisi",
                   "Segment", "Ort. Yorum", 5, 1, "G20")
        print("  [✓] Grafik 5: Fiyat segmenti analizi")

        ws = wb["G6_Kat_Yorum"]
        baslik_formatla(ws, "A9D18E")
        sutun_genislet(ws)
        bar_grafik(ws, "Kategoriye Gore Ortalama Yorum Sayisi",
                   "Kategori", "Ort. Yorum", 2, 1, "D2")
        print("  [✓] Grafik 6: Kategori bazli ortalama yorum sayisi")

    except KeyError as e:
        print(f"HATA: Sayfa bulunamadi: {e}")
    except Exception as e:
        print(f"HATA: Grafik eklenirken hata: {e}")
    finally:
        try:
            wb.save(excel_yol)
            print(f"\nTAMAMLANDI! Rapor kaydedildi: {excel_yol}")
        except PermissionError:
            print(f"HATA: '{excel_yol}' baska bir program tarafindan acik.")
        except Exception as e:
            print(f"HATA: Dosya kaydedilemedi: {e}")